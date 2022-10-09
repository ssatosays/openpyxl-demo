from io import BytesIO

from flask import Flask, render_template, send_file
from openpyxl import load_workbook

app = Flask(__name__)


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/download')
def download():
    wb = load_workbook('template.xlsx')
    ws = wb.active

    ws['B2'].value = '#'
    ws['B3'].value = 'Tokyo'
    ws['B4'].value = 'Kanagawa'
    ws['B5'].value = 'Osaka'
    ws['B6'].value = 'Aichi'
    ws['B7'].value = 'Saitama'

    ws['C2'].value = 'Population'
    ws['C3'].value = '14,011,487'
    ws['C4'].value = '9,236,428'
    ws['C5'].value = '8,807,279'
    ws['C6'].value = '7,516,008'
    ws['C7'].value = '7,340,945'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    wb.close()

    return send_file(output, download_name='template.xlsx', as_attachment=True)


if __name__ == '__main__':
    app.run(debug=True)
